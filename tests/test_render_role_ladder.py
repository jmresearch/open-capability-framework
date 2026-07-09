import csv
import os
import shutil
import subprocess
import sys

import pytest
import yaml

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
FIXTURE_ROLE = {
    "role": "Fixture Manager", "slug": "zz-fixture", "variant": "manager",
    "minted": "2026-07-09", "derived": True,
    "levels": [
        {"code": "M1", "title": "Manager", "scope": "one team", "focus": "one team"},
        {"code": "M2", "title": "Senior Manager", "scope": "a critical team", "focus": "critical team"},
    ],
    "competencies": [
        {"theme": "Onboarding & team formation", "capability": "EM-13",
         "key_area": "People", "key_attribute": "Team Composition",
         "anchor": "tuckman1965",
         "mappings": {
             "M1": {"p": "P2", "sources": ["bauer2010"], "why": "Line manager runs onboarding independently."},
             "M2": {"p": "P3", "sources": ["tuckman1965"], "why": "Reads and intervenes on formation stage."},
         },
         "evidence": {"M1": "New hire ships a real change in their first sprint."}},
        {"theme": "Delegation & empowerment", "capability": "EM-14",
         "key_area": "People", "key_attribute": "Team Composition",
         "anchor": "tuckman1965",
         "mappings": {
             "M1": {"p": "P2", "sources": ["gitlab-em-jd"], "why": "Delegates outcomes within one team."},
             "M2": {"p": "P3", "sources": ["gitlab-em-jd"], "why": "Hands ownership of critical work."},
         }},
    ],
}


@pytest.fixture()
def fixture_role():
    d = os.path.join(ROOT, "roles", "zz-fixture")
    os.makedirs(d, exist_ok=True)
    with open(os.path.join(d, "role.yaml"), "w") as f:
        yaml.safe_dump(FIXTURE_ROLE, f, sort_keys=False)
    yield "zz-fixture"
    shutil.rmtree(d)


def catalog_bar(cap_id, p):
    scale = {r["level"]: r["name"].lower() for r in
             csv.DictReader(open(os.path.join(ROOT, "data", "proficiency_scale.csv")))}
    for r in csv.DictReader(open(os.path.join(ROOT, "data", "capabilities.csv"), encoding="utf-8-sig")):
        if r["id"] == cap_id:
            return r[f"{p}_{scale[p]}"].strip()
    raise KeyError(cap_id)


def test_renders_bar_verbatim_plus_evidence_and_scope(fixture_role):
    from render_role_ladder import render_role
    assert render_role(fixture_role) is True
    md = open(os.path.join(ROOT, "roles", fixture_role, "ladder.md")).read()
    bar = catalog_bar("EM-13", "P2")
    assert bar in md
    assert "*Evidenced by:* New hire ships a real change in their first sprint." in md
    assert "*Why this level:* Line manager runs onboarding independently." in md
    assert "Tuckman, Developmental Sequence in Small Groups" in md  # anchor + sources resolved
    rows = list(csv.reader(open(os.path.join(ROOT, "roles", fixture_role, "ladder.csv"))))
    skill = [r for r in rows if r and r[0] == "skill" and r[3] == "Onboarding & team formation"][0]
    assert skill[5] == f"Depth: {bar} Evidenced by: New hire ships a real change in their first sprint. Scope: one team."
    em14 = [r for r in rows if r and r[0] == "skill" and r[3] == "Delegation & empowerment"][0]
    assert em14[5] == f"Depth: {catalog_bar('EM-14', 'P2')} Scope: one team."  # no evidence clause


def test_generated_csv_passes_ladder_validator(fixture_role):
    from render_role_ladder import render_role
    render_role(fixture_role)
    res = subprocess.run(
        [sys.executable, os.path.join(ROOT, "skills", "career-ladder", "scripts", "validate_csv.py"),
         os.path.join(ROOT, "roles", fixture_role, "ladder.csv"), "--manager"],
        capture_output=True, text=True)
    assert res.returncode == 0, res.stdout


def test_skips_non_derived_and_rejects_proposed(fixture_role):
    from render_role_ladder import render_role
    path = os.path.join(ROOT, "roles", fixture_role, "role.yaml")
    role = yaml.safe_load(open(path))
    role["derived"] = False
    yaml.safe_dump(role, open(path, "w"), sort_keys=False)
    assert render_role(fixture_role) is False
    role["derived"] = True
    role["competencies"][0]["capability"] = "proposed"
    yaml.safe_dump(role, open(path, "w"), sort_keys=False)
    with pytest.raises(SystemExit):
        render_role(fixture_role)


def test_renders_none_evidence_as_bar_and_scope_only(fixture_role):
    from render_role_ladder import render_role
    path = os.path.join(ROOT, "roles", fixture_role, "role.yaml")
    role = yaml.safe_load(open(path))
    role["competencies"][0]["evidence"] = {"M1": None}  # `M1:` with no value in yaml
    yaml.safe_dump(role, open(path, "w"), sort_keys=False)
    assert render_role(fixture_role) is True
    rows = list(csv.reader(open(os.path.join(ROOT, "roles", fixture_role, "ladder.csv"))))
    skill = [r for r in rows if r and r[0] == "skill" and r[3] == "Onboarding & team formation"][0]
    bar = catalog_bar("EM-13", "P2")
    assert skill[5] == f"Depth: {bar} Scope: one team."


def test_xlsx_has_theory_columns_and_sources_tab(fixture_role, tmp_path):
    from render_role_ladder import render_role
    render_role(fixture_role)
    subprocess.run(
        [sys.executable, os.path.join(ROOT, "scripts", "render_role_xlsx.py"), fixture_role],
        check=True, capture_output=True, text=True)
    import openpyxl
    wb = openpyxl.load_workbook(os.path.join(ROOT, "roles", fixture_role, "ladder.xlsx"))
    assert "Sources & Theory" in wb.sheetnames
    assert "Korn Ferry Crosswalk" not in wb.sheetnames
    matrix = wb["Competency Matrix"]
    headers = [c.value for c in matrix[1]]
    assert "What it covers" in headers and "Theory anchor & why" in headers
    src = wb["Sources & Theory"]
    assert [c.value for c in src[1]] == ["Framework", "Source", "Grounds", "Link"]
    cells = [str(c.value) for row in src.iter_rows() for c in row if c.value]
    assert any("Tuckman" in v for v in cells)
    os.remove(os.path.join(ROOT, "roles", fixture_role, "ladder.xlsx"))
