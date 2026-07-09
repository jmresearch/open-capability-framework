#!/usr/bin/env python3
"""Render roles/<slug>/ladder.xlsx from a role record (role.yaml + ladder.md + ladder.csv).

Sheets: Overview, Competency Matrix (OCF ids hyperlinked to the GitHub catalog),
Rating Template (self/peer/manager inputs with IFERROR-guarded formulas),
Sources & Theory.

Re-runnable. Run with:
    uv run --with openpyxl --with pyyaml python scripts/render_role_xlsx.py [slug ...]
"""

import csv
import os
import re
import sys

import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
def _discover_roles():
    _base = os.path.join(ROOT, "roles")
    return tuple(sorted(d for d in os.listdir(_base)
                        if os.path.isfile(os.path.join(_base, d, "role.yaml"))))

DEFAULT_ROLES = _discover_roles()
REPO = "https://github.com/jmresearch/open-capability-framework"

FONT = "Arial"
SIZE = 10
BODY = Font(name=FONT, size=SIZE)
BOLD = Font(name=FONT, size=SIZE, bold=True)
INPUT = Font(name=FONT, size=SIZE, color="FF0000CC")  # blue = input convention
LINK = Font(name=FONT, size=SIZE, color="FF0563C1", underline="single")
WRAP = Alignment(wrap_text=True, vertical="top")
TOP = Alignment(vertical="top")


def set_widths(ws, widths):
    for j, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w


def load_sourcing():
    def rd(name):
        with open(os.path.join(ROOT, "data", name), encoding="utf-8-sig", newline="") as f:
            return list(csv.DictReader(f))
    caps = {r["id"].strip(): r for r in rd("capabilities.csv")}
    bib = {r["key"].strip(): r for r in rd("bibliography.csv")}
    return caps, bib


def load_role(slug):
    role_dir = os.path.join(ROOT, "roles", slug)
    with open(os.path.join(role_dir, "role.yaml"), encoding="utf-8") as f:
        role = yaml.safe_load(f)
    with open(os.path.join(role_dir, "ladder.md"), encoding="utf-8") as f:
        ladder_md = f.read()
    with open(os.path.join(role_dir, "ladder.csv"), encoding="utf-8-sig", newline="") as f:
        rows = list(csv.reader(f))
    return role, ladder_md, rows


def skill_cells(rows):
    """Map (key_area, key_attribute, theme) -> per-level cell prose from ladder.csv."""
    header = rows[0]
    first_level = 5 if header[4].strip() == "capability" else 4
    n_levels = len(header) - first_level
    out = {}
    for row in rows[1:]:
        if row and row[0].strip() == "skill":
            cells = (row[first_level:] + [""] * n_levels)[:n_levels]
            out[(row[1], row[2], row[3])] = cells
    return out


def ocf_link(comp):
    """(display text, url) for a competency's OCF reference."""
    cap = comp.get("capability")
    if cap == "proposed":
        fname = os.path.basename(comp["proposal_ref"])
        return "proposed", f"{REPO}/blob/main/contrib/{fname}"
    return cap, f"{REPO}/blob/main/data/capabilities.md#{cap.lower()}"


def sources_list(ladder_md):
    m = re.search(r"^## Sources\s*$(.*)", ladder_md, re.M | re.S)
    if not m:
        return []
    seen, out = set(), []
    for line in m.group(1).split("\n"):
        line = line.strip()
        if line.startswith("- "):
            src = line[2:].strip()
            if src not in seen:
                seen.add(src)
                out.append(src)
    return out


# ------------------------------------------------------------------ sheets --

def sheet_overview(wb, role):
    ws = wb.active
    ws.title = "Overview"
    minted = str(role.get("minted", ""))
    meta = [("Role", role["role"]), ("Slug", role["slug"]),
            ("Variant", role["variant"]), ("Minted", minted)]
    for label, value in meta:
        ws.append([label, value])
        ws.cell(row=ws.max_row, column=1).font = BOLD
        ws.cell(row=ws.max_row, column=2).font = BODY
    ws.append([])
    ws.append(["Level", "Title", "Scope", "Focus"])
    for c in ws[ws.max_row]:
        c.font = BOLD
    for lv in role["levels"]:
        ws.append([lv["code"], lv["title"], lv["scope"], lv["focus"]])
        for c in ws[ws.max_row]:
            c.font = BODY
            c.alignment = WRAP if c.column >= 3 else TOP
    ws.append([])
    note = ("How to read: each row of the Competency Matrix is one competency, mapped to an "
            "Open Capability Framework capability (OCF ID column — click through for the "
            "catalog entry and its P1-P6 profile). Level columns describe observable "
            "behavior at each level; use the Rating Template for self/peer/manager scoring.")
    ws.append([note])
    ws.cell(row=ws.max_row, column=1).font = BODY
    ws.cell(row=ws.max_row, column=1).alignment = WRAP
    ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=4)
    ws.row_dimensions[ws.max_row].height = 60
    ws.append([])
    ws.append(["Framework repository", REPO])
    ws.cell(row=ws.max_row, column=1).font = BOLD
    link = ws.cell(row=ws.max_row, column=2)
    link.hyperlink = REPO
    link.font = LINK
    set_widths(ws, [22, 36, 34, 60])


def sheet_matrix(wb, role, cells_by_triple, caps, bib):
    ws = wb.create_sheet("Competency Matrix")
    codes = [lv["code"] for lv in role["levels"]]
    ws.append(["Key Area", "Focus Area", "Competency", "OCF ID",
               "What it covers", "Theory anchor & why"] + codes)
    for c in ws[1]:
        c.font = BOLD
    for comp in role["competencies"]:
        triple = (comp["key_area"], comp["key_attribute"], comp["theme"])
        cells = cells_by_triple.get(triple)
        if cells is None:
            raise SystemExit(f"no ladder.csv skill row for {triple}")
        text, url = ocf_link(comp)
        cap_row = caps.get(str(comp.get("capability", "")).strip())
        covers = cap_row["description"] if cap_row else ""
        anchor_key = str(comp.get("anchor", "")).strip()
        anchor = bib[anchor_key]["citation"] if anchor_key in bib else anchor_key
        if comp.get("anchor_why"):
            anchor += f" — {comp['anchor_why']}"
        ws.append([comp["key_area"], comp["key_attribute"], comp["theme"], text,
                   covers, anchor] + cells)
        r = ws.max_row
        for c in ws[r]:
            c.font = BODY
            c.alignment = WRAP
        idc = ws.cell(row=r, column=4)
        idc.hyperlink = url
        idc.font = LINK
    set_widths(ws, [18, 18, 24, 12, 45, 45] + [55] * len(codes))
    ws.freeze_panes = "A2"


def sheet_rating(wb, role):
    ws = wb.create_sheet("Rating Template")
    codes = [lv["code"] for lv in role["levels"]]
    # default target: terminal level for ICs (marked in scope), else first level
    default = codes[0]
    for lv in role["levels"]:
        if "terminal" in str(lv.get("scope", "")).lower():
            default = lv["code"]

    ws["A1"] = "Target level"
    ws["A1"].font = BOLD
    ws["B1"] = default
    ws["B1"].font = INPUT
    ws["D1"] = f"Enter one of: {', '.join(codes)}. Blue cells are inputs (1-6, the P number)."
    ws["D1"].font = BODY

    headers = ["Key Area", "Competency", "Target P", "Self", "Peer", "Manager", "Avg", "Gap"]
    hdr_row = 3
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=hdr_row, column=j, value=h)
        c.font = BOLD
    # hidden per-level P-number block: J..O headers = level codes
    first_hidden = 10  # column J
    for j, code in enumerate(codes):
        c = ws.cell(row=hdr_row, column=first_hidden + j, value=code)
        c.font = BOLD
    hcol0 = get_column_letter(first_hidden)
    hcol1 = get_column_letter(first_hidden + len(codes) - 1)

    r = hdr_row
    for comp in role["competencies"]:
        r += 1
        ws.cell(row=r, column=1, value=comp["key_area"]).font = BODY
        ws.cell(row=r, column=2, value=comp["theme"]).font = BODY
        tgt = ws.cell(row=r, column=3)
        tgt.value = (f'=IFERROR(INDEX({hcol0}{r}:{hcol1}{r},'
                     f'MATCH($B$1,${hcol0}${hdr_row}:${hcol1}${hdr_row},0)),"")')
        tgt.font = BODY
        for j in (4, 5, 6):  # Self / Peer / Manager inputs
            ws.cell(row=r, column=j).font = INPUT
        avg = ws.cell(row=r, column=7)
        avg.value = f'=IFERROR(AVERAGE(D{r}:F{r}),"")'
        avg.font = BODY
        gap = ws.cell(row=r, column=8)
        gap.value = f'=IFERROR(G{r}-C{r},"")'
        gap.font = BODY
        for j, code in enumerate(codes):
            p = comp["mappings"][code]["p"]
            ws.cell(row=r, column=first_hidden + j, value=int(p.lstrip("P"))).font = BODY

    for j in range(first_hidden, first_hidden + len(codes)):
        ws.column_dimensions[get_column_letter(j)].hidden = True
    set_widths(ws, [26, 34, 9, 7, 7, 9, 7, 7])
    ws.freeze_panes = f"A{hdr_row + 1}"


def sheet_sources(wb, role, bib, ladder_md):
    ws = wb.create_sheet("Sources & Theory")
    ws.append(["Framework", "Source", "Grounds", "Link"])
    for c in ws[1]:
        c.font = BOLD
    seen = set()
    for comp in role["competencies"]:
        key = str(comp.get("anchor", "")).strip()
        row = bib.get(key)
        grounds = comp["theme"]
        if row:
            if key in seen:
                continue
            seen.add(key)
            ws.append([row["notes"] or comp["theme"], row["citation"], grounds, row["link"]])
        else:
            ws.append([comp["theme"], key or "(unsourced)", grounds, ""])
        for c in ws[ws.max_row]:
            c.font = BODY
            c.alignment = WRAP
    for src in sources_list(ladder_md):          # legacy free-text Sources section, if any
        ws.append(["", src, "", ""])
        for c in ws[ws.max_row]:
            c.font = BODY
            c.alignment = WRAP
    set_widths(ws, [30, 70, 40, 45])
    ws.freeze_panes = "A2"


def render(slug):
    role, ladder_md, rows = load_role(slug)
    caps, bib = load_sourcing()
    wb = Workbook()
    sheet_overview(wb, role)
    sheet_matrix(wb, role, skill_cells(rows), caps, bib)
    sheet_rating(wb, role)
    sheet_sources(wb, role, bib, ladder_md)
    out = os.path.join(ROOT, "roles", slug, "ladder.xlsx")
    wb.save(out)
    print(f"wrote roles/{slug}/ladder.xlsx "
          f"({len(role['competencies'])} competencies, {len(role['levels'])} levels)")


def main():
    for slug in sys.argv[1:] or DEFAULT_ROLES:
        render(slug)


if __name__ == "__main__":
    main()
