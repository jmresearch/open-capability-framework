#!/usr/bin/env python3
"""Render the human-usable catalog layer from data/*.csv.

Writes (all under data/):
    proficiency_scale.md   — the P1-P6 rubric with stable anchors (#p1..#p6)
    scope_levels.md        — the S1-S6 role-scope axis with anchors (#s1..#s6)
    capabilities.md        — the full capability catalog, one section per
                             capability with a stable anchor (#<id-lowercase>)
                             and all six proficiency profiles linked to the scale
    capabilities.xlsx      — the same data as a formatted workbook

Re-runnable: regenerates every output from the CSVs each time.
Run with:  uv run --with openpyxl python scripts/render_catalog.py
"""

import csv
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA = os.path.join(ROOT, "data")

FONT_NAME = "Arial"
FONT_SIZE = 10


def read_csv(name):
    with open(os.path.join(DATA, name), encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def load_sourcing():
    """(bibliography by key, capability_levels by (capability_id, level))."""
    bib = {r["key"].strip(): r for r in read_csv("bibliography.csv")}
    levels = {(r["capability_id"].strip(), r["level"].strip()): r
              for r in read_csv("capability_levels.csv")}
    return bib, levels


def write_text(name, text):
    with open(os.path.join(DATA, name), "w", encoding="utf-8", newline="\n") as f:
        f.write(text)
    print(f"wrote data/{name}")


# ---------------------------------------------------------------- markdown --

def render_proficiency_md(scale):
    out = [
        "# Proficiency Scale (P1–P6)",
        "",
        "The universal capability-side proficiency rubric: every capability in "
        "[the catalog](capabilities.md) is leveled against these six Dreyfus-derived "
        "bands, independent of any role's scope (see [scope levels](scope_levels.md)).",
        "",
    ]
    for row in scale:
        lvl = row["level"].strip()
        out += [
            f'<a id="{lvl.lower()}"></a>',
            f"### {lvl} — {row['name'].strip()}",
            "",
            f"*Dreyfus analog:* {row['dreyfus_analog'].strip()}",
            "",
            row["definition"].strip(),
            "",
        ]
    return "\n".join(out)


def render_scope_md(scopes):
    out = [
        "# Scope Levels (S1–S6)",
        "",
        "The role-side scope axis: how broad a role's remit is, from a single task "
        "to the whole company. Scope is a property of the **role**, not the "
        "capability — capability depth is measured on the "
        "[P1–P6 proficiency scale](proficiency_scale.md).",
        "",
    ]
    for row in scopes:
        s = row["scope"].strip()
        out += [
            f'<a id="{s.lower()}"></a>',
            f"### {s} — {row['blast_radius'].strip()}",
            "",
            f"*Blast radius:* {row['blast_radius'].strip()}",
            "",
            f"*SFIA level of responsibility:* {row['sfia_lor'].strip()}",
            "",
            f"*Typical IC title:* {row['typical_ic_title'].strip()}",
            "",
            f"*Google / Meta anchor:* {row['google_meta'].strip()}",
            "",
        ]
    return "\n".join(out)


def render_capabilities_md(caps, scale, domains, bib, levels):
    p_cols = []  # [(csv_column, level, name)]
    for row in scale:
        lvl, name = row["level"].strip(), row["name"].strip()
        p_cols.append((f"{lvl}_{name.lower()}", lvl, name))
    prefix_by_domain = {d["domain"].strip(): d["prefix"].strip() for d in domains}

    out = [
        "# Capability Catalog",
        "",
        f"The full Open Capability Framework catalog — {len(caps)} capabilities "
        f"across {len(domains)} domains, grouped Segment → Domain → Focus Area. Each "
        "capability carries a six-point behavioral profile on the capability-side "
        "[P1–P6 proficiency scale](proficiency_scale.md). Note the two-axis "
        "split: proficiency (P1–P6) is intrinsic to the capability and "
        "measures *how well*; scope ([S1–S6](scope_levels.md)) is a property "
        "of the **role** and measures *how broad*. Role records under `roles/` "
        "combine the two.",
        "",
        "Each capability has a stable anchor equal to its lowercased id "
        "(e.g. `#em-03`) so role ladders can deep-link here.",
        "",
    ]
    seg = dom = focus = None
    for row in caps:
        if row["segment"] != seg:
            seg = row["segment"]
            dom = focus = None
            out += [f"## {seg}", ""]
        if row["domain"] != dom:
            dom = row["domain"]
            focus = None
            prefix = prefix_by_domain.get(dom.strip(), "")
            out += [f"### {dom} ({prefix})", ""]
        if row["focus_area"] != focus:
            focus = row["focus_area"]
            out += [f"#### {focus}", ""]
        cid = row["id"].strip()
        out += [
            f'<a id="{cid.lower()}"></a>',
            f"##### {cid} — {row['capability'].strip()}",
            "",
            f"*Type:* {row['type'].strip()} — {row['description'].strip()}",
            "",
        ]
        for col, lvl, name in p_cols:
            line = (f"- **[{lvl} — {name}](proficiency_scale.md#{lvl.lower()}):** "
                    f"{row[col].strip()}")
            src = levels.get((cid, lvl))
            if src:
                cites = "; ".join(
                    bib[k]["citation"] for k in src["source_keys"].split(";") if k.strip()
                )
                line += f"\n  — *Why this level:* {src['why'].strip()} *Sources:* {cites}."
            out.append(line)
        out.append("")
    return "\n".join(out)


# -------------------------------------------------------------------- xlsx --

def style_sheet(ws, widths, wrap_cols=()):
    header_font = Font(name=FONT_NAME, size=FONT_SIZE, bold=True)
    body_font = Font(name=FONT_NAME, size=FONT_SIZE)
    wrap = Alignment(wrap_text=True, vertical="top")
    top = Alignment(vertical="top")
    for j, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w
    for i, row in enumerate(ws.iter_rows(), start=1):
        for j, cell in enumerate(row, start=1):
            cell.font = header_font if i == 1 else body_font
            cell.alignment = wrap if (i > 1 and j in wrap_cols) else top
    ws.freeze_panes = "A2"


def render_xlsx(caps, scale, scopes, domains, bib):
    wb = Workbook()

    ws = wb.active
    ws.title = "Catalog"
    ws.append(["ID", "Segment", "Domain", "Focus Area", "Capability", "Type",
               "Description", "P1", "P2", "P3", "P4", "P5", "P6"])
    p_cols = [f"{r['level'].strip()}_{r['name'].strip().lower()}" for r in scale]
    for row in caps:
        ws.append([row["id"], row["segment"], row["domain"], row["focus_area"],
                   row["capability"], row["type"], row["description"]]
                  + [row[c] for c in p_cols])
    style_sheet(ws, [9, 24, 26, 26, 30, 11, 45, 40, 40, 40, 40, 40, 40],
                wrap_cols={7, 8, 9, 10, 11, 12, 13})

    ws = wb.create_sheet("Proficiency Scale")
    ws.append(["Level", "Name", "Dreyfus Analog", "Definition"])
    for r in scale:
        ws.append([r["level"], r["name"], r["dreyfus_analog"], r["definition"]])
    style_sheet(ws, [8, 14, 18, 90], wrap_cols={4})

    ws = wb.create_sheet("Scope Levels")
    ws.append(["Scope", "Blast Radius", "SFIA LoR", "Typical IC Title", "Google / Meta"])
    for r in scopes:
        ws.append([r["scope"], r["blast_radius"], r["sfia_lor"],
                   r["typical_ic_title"], r["google_meta"]])
    style_sheet(ws, [8, 14, 10, 28, 14])

    ws = wb.create_sheet("Domains")
    ws.append(["Domain", "Segment", "Prefix", "Focus Areas", "Capabilities"])
    for r in domains:
        ws.append([r["domain"], r["segment"], r["prefix"],
                   int(r["focus_areas"]), int(r["capabilities"])])
    style_sheet(ws, [42, 32, 9, 12, 12])

    ws = wb.create_sheet("Sources & Theory")
    ws.append(["Key", "Type", "Citation", "Link", "Notes"])
    for r in bib.values():
        ws.append([r["key"], r["type"], r["citation"], r["link"], r["notes"]])
    style_sheet(ws, [18, 14, 70, 40, 40], wrap_cols={3, 5})

    path = os.path.join(DATA, "capabilities.xlsx")
    wb.save(path)
    print("wrote data/capabilities.xlsx")


def main():
    caps = read_csv("capabilities.csv")
    scale = read_csv("proficiency_scale.csv")
    scopes = read_csv("scope_levels.csv")
    domains = read_csv("domains.csv")

    bib, levels = load_sourcing()

    write_text("proficiency_scale.md", render_proficiency_md(scale))
    write_text("scope_levels.md", render_scope_md(scopes))
    write_text("capabilities.md", render_capabilities_md(caps, scale, domains, bib, levels))
    render_xlsx(caps, scale, scopes, domains, bib)
    print(f"rendered {len(caps)} capabilities")


if __name__ == "__main__":
    main()
